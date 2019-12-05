#_*_coding=utf-8_*_

'''
v3.0 考虑到TXT无法进行指定行输出，使用Excel尝试
v4.0 因为主用for语句无法和嵌套的for语句共享循环次数，现在while为主循环
20190514 补充了VRF description部分命令
20190523补充vlan和vlan名字以及list部分信息录入
'''

'''
已知的，但是还没解决的问题：
1.span的源接口可能是vpws
'''



from openpyxl import load_workbook
import re
import easygui as gui

#abcdefgh
#opqrst
#uvwxyz

a = 1   #每次在读到第一行的时候到下一行
b = 0
c = 0   #VRF的for循环计数

d = 0  #用于L2port的for循环计数
e = 1  #用于L2port的行数计数

f = 0  #用于vlan的for循环计数
g = 1  #用于vlan的行数计数

vlan_name_count = 1 #循环vlan和vlan名字的计数器
vlan_list_count = 1 #循环vlan list的计数器

o = 0  #用于lacp的for循环计数
p = 1  #用于lacp的行数计数

q = 0  #用于ospf的for循环函数 
r = 1  #用于ospf的行数计数

s = 0  #用于vrrp的for循环函数
t = 1  #用于vrrp的行计数

u = 0  #用于span的for循环计数
v = 1  #用于span的行计数

n = 1 #原始配置的行
i = 1 #原始配置的列

count = 0 #内置的计数器
#cell_list  #存储被划分的单元格生成的数组
#cell_temp  列表操作的过度



vrf_count = 0

vrf_module = re.compile("!<vrf>")
vrf_module_end = re.compile("!</vrf>")
vrf_name   = re.compile('ip vrf')
vrf_rd     = re.compile('rd')
vrf_rt_imp = re.compile('route-target import')
vrf_rt_exp = re.compile('route-target export')
vrf_des = re.compile('description')


l2port_count = 0

l2port_module = re.compile('!<if-intf>')
l2port_name = re.compile('interface')
l2port_description = re.compile('description')
l2port_mtu = re.compile('mtu')
l2port_status = re.compile('no shutdown')
l3port_ip = re.compile('ip address')
l3port_vrf = re.compile('ip vrf forwarding')
l3port_mtu = re.compile('ip mtu')
#smartgroup口已经被覆盖
#三层口的接口描述和
#trunktmp
#存储原来的trunk vlan数据
#trunkcell
#用于最终写入单元格的
l2portdefine = re.compile(r'\w*gei\w*')#这个是用来查找匹配的
smartport = re.compile(r'\w*smartgroup\w*')
#l2l3port 用来判断是二层还是三层口
l2port_module_end = re.compile('!</if-intf>')

vlan_count = 0

vlan_module = re.compile('!<switchvlan>')
vlan_num = re.compile(r'^vlan\s*\w')
vlan_name = re.compile(r'^name\s*\w')
vlan_list = re.compile(r'^list\s*\w')
vlan_port_name = re.compile('interface')
vlan_trunk = re.compile('switchport mode trunk')
vlan_trunk_vlan = re.compile('switchport trunk vlan')
vlan_trunk_native = re.compile('switchport trunk native vlan')
vlan_access = re.compile('switchport access vlan')
vlan_hybrid = re.compile("switchport mode hybrid")
vlan_hybrid_tagvlan = re.compile(r"\w*\stag*\w")
vlan_hybrid_untagvlan = re.compile(r"\w*\suntag*\w")
vlan_hybrid_native = re.compile("switchport hybrid native vlan")

vlan_module_end = re.compile('!</switchvlan>')

lacp_count = 0

lacp_module = re.compile('!<lacp>')
lacp_name = re.compile('interface')
lacp_mode = re.compile('lacp mode')
lacp_num = re.compile('smartgroup')
lacp_load_balance = re.compile('lacp load-balance ')
lacp_module_end = re.compile('!</lacp>')

ospf_count = 0

ospf_module = re.compile('!<ospfv2>')
ospf_process = re.compile('router ospf')
ospf_area = re.compile('area')
ospf_network = re.compile('network')
ospf_te = re.compile('mpls traffic-eng area')
ospf_red = re.compile('redistribute')
ospf_module_end = re.compile('!</ospfv2>')

'''
ospf目前只做ospfV2的，但是配置比较稀少，此外，还需要搜集不同VRF的。
mpls l2vpn 是一套的 ldp是需要的，l2vpn是需要的，二层功能启用的命令。mpls interface 怎么样？？
route-map prefix-list acl as-path 
'''

bgp_count = 0
'''
bgp_module
bgp_as
bgp_neighbor
bgp_
'''

vrrp_count = 0
vrrp_module = re.compile('!<vrrp>')
vrrp_intf = re.compile('interface')
vrrp_number = re.compile(r'^vrrp\s\d{1,3}\sipv4\s') 
vrrp_priority = re.compile(r'^vrrp\s\d{1,3}\spriority')
vrrp_module_end = re.compile('!</vrrp>')

span_count = 0
span_module = re.compile('!<monitor>')
span_session_num = re.compile('span session')
span_dest_intf = re.compile('default destination interface')
span_source = re.compile('span apply session')
span_end_module = re.compile('!</monitor>')

ldp_count = 0
ldp_module = re.compile('!<ldp>')
ldp_instace = re.compile('mpls ldp instance')
ldp_interface = re.compile(r'^interface')
ldp_routerid = re.compile('router-id')
ldp_target = re.compile('target-session')

ldp_end_module = re.compile('!</ldp>')

vpls_count = 0
vpls_module = re.compile('!<l2vpn>')
#在转换为脚本输出的时候需要自动添加mpsl l2vpn enable命令

vpls_name = re.compile(r'^vpls')
vpls_acpoint = re.compile('access-point')
vpls_acparams = re.compile('access-params')
vpls_pw = re.compile('pseudo-wire')
vpls_pw_neighbor = re.compile('neighbour')
vpls_pw_tunnelpolicy = re.compile('tunnel-policy')


vpls_end_module = re.compile('!</l2vpn>')
#还有一些的功能性配置，可以的话补充一下




#打开保存配置的Excel
config = load_workbook('configuration.xlsx')
#con_sheet = config.get_sheet_by_name('configure') 这种会产生告警
con_sheet = config['configure']

nrows = con_sheet.max_row #行
ncols = con_sheet.max_column #列


newwb = load_workbook('test.xlsx')
#打开存储Excel
#sh = newwb.get_sheet_by_name('vrrp')
#示例：按照名字打开Excel
sh = newwb['vrrp']



while n < nrows:
	#必须从0开始的
	cell = con_sheet.cell(n,1).value
	if cell == None:
		continue
	else:
		cell = cell.lstrip()
	#print(cell)
	n = n + 1
	#openpyxl从1开始计数，但是首行有表头，应该从2开始
	#匹配模块进入本模块的sheet
	if vrf_module.match(cell):
		sheet_vrf = newwb.get_sheet_by_name('vrf')
		vrf_count = n
		#调用整体的行数进入自己的循环
		print("开始VRF信息的写入！！！")
		#print(vrf_count)
		for c in range(vrf_count,nrows):
			#因为每次进入一个表都是单独的新表
			#每次需要从第二行开始写
			#列数也需要进行重置

			cell_vrf = con_sheet.cell(c,1).value
			cell_vrf = cell_vrf.lstrip()
			#print(c)
			#print(cell_vrf)
			if vrf_name.match(cell_vrf):
				#逐行匹配，查找匹配项
				cell_list = cell_vrf.split(' ')
				cell_vrf = cell_list[2]
				#print(cell_vrf)
				a = a + 1
				#进入下一行，这里跳行的话，避免了没有配置rt的尴尬情况
				sheet_vrf.cell(a,1).value = cell_vrf
				#还是固定列比较好，风险在于要严格区分，不能有重复匹配的
				#行数因为不知道有多少，需要循环指定
				#print(b)
			elif vrf_des.match(cell_vrf):
				cell_list = cell_vrf.split(' ')
				cell_temp = cell_list.pop(0)
				cell_vrf = " ".join(cell_list)
				sheet_vrf.cell(a,5).value = cell_vrf

			elif vrf_rd.match(cell_vrf):
				cell_list = cell_vrf.split(' ')
				cell_vrf = cell_list[1]
				sheet_vrf.cell(a,2).value = cell_vrf
				#print(b)
			elif vrf_rt_imp.match(cell_vrf):
				cell_list = cell_vrf.split(' ')
				cell_vrf = cell_list[2]
				sheet_vrf.cell(a,3).value = cell_vrf
				#print(b)
			elif vrf_rt_exp.match(cell_vrf):
				cell_list = cell_vrf.split(' ')
				cell_vrf = cell_list[2]
				sheet_vrf.cell(a,4).value = cell_vrf
				#print(a)
				#print(b)
			elif vrf_module_end.match(cell_vrf):
				#跳过本行，跳出本模块，进行下个模块的匹配
				n = c + 1
				#对n重新赋值，下次循环就从后面一行开始了
				#归还循环的值到n，这样n可以继续使用
				print(n)
				print("VRF 部分已经完成输出，Excel写入成功！！！")
				break
			else:
				continue
				#如果是$换行符，则继续循环
				#只准成功，不能失败，不然n的值就不能修改了
		#print(n)
	elif l2port_module.match(cell):
		sheet_l2port = newwb['l2port']
		l2port_count = n
		print('开始进行二层三层接口信息的写入！！！')
		print(l2port_count)
		for d in range(l2port_count,nrows):
			cell_l2port = con_sheet.cell(d,1).value
			cell_l2port = cell_l2port.lstrip()
			if l2port_name.match(cell_l2port):
				e = e + 1
				cell_list = cell_l2port.split(' ')
				cell_l2port = cell_list[1]
				sheet_l2port.cell(e,1).value = cell_l2port
				#如何判断是二层三层口？
				if l2portdefine.findall(cell_l2port):
					sheet_l2port.cell(e,2).value = 'L2'
				elif smartport.findall(cell_l2port):
					sheet_l2port.cell(e,2).value = 'smartgroup口'
				else:
					sheet_l2port.cell(e,2).value = 'L3'
			elif l2port_status.match(cell_l2port):
				sheet_l2port.cell(e,3).value = cell_l2port
			elif l2port_description.match(cell_l2port):
				cell_list = cell_l2port.split(' ')
				cell_temp = cell_list.pop(0)
				cell_l2port = " ".join(cell_list)
				#description需要将头部的去掉，因为描述里面可以加空格
				#再把原先的空格加上，变成描述
				sheet_l2port.cell(e,6).value = cell_l2port
			elif l2port_mtu.match(cell_l2port):
				cell_list = cell_l2port.split(' ')
				cell_l2port = cell_list[1]
				sheet_l2port.cell(e,8).value = cell_l2port
			elif l3port_ip.match(cell_l2port):
				cell_list = cell_l2port.split(' ')
				cell_l2port = cell_list[2] + ' ' + cell_list[3]
				#20190505 最好还是把掩码带上
				sheet_l2port.cell(e,4).value = cell_l2port
			elif l3port_mtu.match(cell_l2port):
				cell_list = cell_l2port.split(' ')
				cell_l2port = cell_list[2]
				sheet_l2port.cell(e,9).value = cell_l2port
			elif l3port_vrf.match(cell_l2port):
				cell_list = cell_l2port.split(' ')
				cell_l2port = cell_list[3]
				sheet_l2port.cell(e,5).value = cell_l2port
			elif l2port_module_end.match(cell_l2port):
				n = d + 1
				print('二三层接口信息已经完成！！！')
				break
			else:
				continue
	elif vlan_module.match(cell):
		sheet_vlan = newwb['vlan']
		vlan_count = n
		print("开始VLAN信息的写入！！！")
		print(vlan_count)
		for f in range(vlan_count,nrows):
			cell_vlan = con_sheet.cell(f,1).value
			cell_vlan = cell_vlan.lstrip()
			if vlan_port_name.match(cell_vlan):
				g = g + 1
				cell_list = cell_vlan.split(' ')
				cell_vlan = cell_list[1]
				sheet_vlan.cell(g,1).value = cell_vlan
			elif vlan_trunk.match(cell_vlan):
				cell_list = cell_vlan.split(' ')
				cell_vlan = cell_list[2]
				sheet_vlan.cell(g,2).value = cell_vlan
			elif vlan_trunk_vlan.match(cell_vlan):
				cell_list = cell_vlan.split(' ')
				cell_vlan = cell_list[3]
				if sheet_vlan.cell(g,4).value == None:
					sheet_vlan.cell(g,4).value = cell_vlan
				else:
					trunktmp = str(sheet_vlan.cell(g,4).value)
					#获取原先的单元格数据
					trunkcell = trunktmp + "," + cell_vlan
					#添加后来的vlan信息
					sheet_vlan.cell(g,4).value = trunkcell
					#将所有的信息写进单元格
			elif vlan_trunk_native.match(cell_vlan):
				cell_list = cell_vlan.split(' ')
				cell_vlan = cell_list[4]
				sheet_vlan.cell(g,5).value = cell_vlan
			elif vlan_hybrid.match(cell_vlan):
				cell_list = cell_vlan.split(' ')
				cell_vlan = cell_list[2]
				sheet_vlan.cell(g,2).value = cell_vlan
			elif vlan_hybrid_tagvlan.findall(cell_vlan):
				#print(cell_vlan)
				cell_list = cell_vlan.split(' ')
				cell_vlan = cell_list[3]
				if sheet_vlan.cell(g,6).value == None:
					sheet_vlan.cell(g,6).value = cell_vlan
				else:
					hybridtmp = sheet_vlan.cell(g,6).value
					hybridcell = hybridtmp + "," + cell_vlan
					sheet_vlan.cell(g,6).value = hybridcell
			elif vlan_hybrid_untagvlan.findall(cell_vlan):
				cell_list = cell_vlan.split(' ')
				cell_vlan = cell_list[3]
				if sheet_vlan.cell(g,7).value == None:
					sheet_vlan.cell(g,7).value = cell_vlan
				else:
					hybridtmp = sheet_vlan.cell(g,7).value
					hybridcell = hybridtmp + ',' + cell_vlan
					sheet_vlan.cell(g,7).value = hybridcell
			elif vlan_access.match(cell_vlan):
				cell_list = cell_vlan.split(' ')
				cell_vlan = cell_list[1]
				sheet_vlan.cell(g,2).value = cell_vlan
				cell_vlan = cell_list[3]
				sheet_vlan.cell(g,3).value = cell_vlan
				#access接口命令都砸一行里面，需要分开输出



			elif vlan_num.match(cell_vlan):
				vlan_name_count = vlan_name_count + 1
				cell_list = cell_vlan.split(' ')
				cell_vlan = cell_list[1]
				sheet_vlan.cell(vlan_name_count,9).value = cell_vlan
			elif vlan_name.match(cell_vlan):
				cell_list = cell_vlan.split(' ')
				cell_temp = cell_list.pop(0)
				cell_vlan = " ".join(cell_list)
				sheet_vlan.cell(vlan_name_count,10).value = cell_vlan
			elif vlan_list.match(cell_vlan):
				vlan_list_count = vlan_list_count + 1
				cell_list = cell_vlan.split(' ')
				cell_vlan = cell_list[1]
				sheet_vlan.cell(vlan_list_count,11).value = cell_vlan


			elif vlan_module_end.match(cell_vlan):
				n = f + 1
				print("VLAN信息写入完成！！！")
				break
			else:
				continue
	elif lacp_module.match(cell):
		sheet_lacp = newwb['lacp']
		lacp_count = n
		print("开始进行LACP信息的写入！！！")
		for o in range(lacp_count,nrows):
			cell_lacp = con_sheet.cell(o,1).value
			cell_lacp = cell_lacp.lstrip()
			if lacp_name.match(cell_lacp):
				p = p + 1
				cell_list = cell_lacp.split(' ')
				cell_lacp = cell_list[1]
				sheet_lacp.cell(p,1).value = cell_lacp
			elif lacp_mode.match(cell_lacp):
				cell_list = cell_lacp.split(' ')
				cell_lacp = cell_list[2]
				sheet_lacp.cell(p,4).value = cell_lacp
			elif lacp_num.match(cell_lacp):
				cell_list = cell_lacp.split(' ')
				cell_lacp = cell_list[1]
				sheet_lacp.cell(p,2).value = cell_lacp
				cell_lacp = cell_list[3]
				sheet_lacp.cell(p,3).value = cell_lacp
			elif lacp_load_balance.match(cell_lacp):
				cell_list = cell_lacp.split(' ')
				cell_lacp = cell_list[2]
				sheet_lacp.cell(p,5).value = cell_lacp
			elif lacp_module_end.match(cell_lacp):
				n = o + 1
				print('LACP信息已经完成写入！！！')
				print(n)
				break
			else:
				continue
				print('n')
	elif ospf_module.match(cell):
		sheet_ospf = newwb['ospf']
		ospf_count = n
		print("开始进行OSPF数据的写入！！！")
		for q in range(ospf_count,nrows):
			cell_ospf = con_sheet.cell(q,1).value
			cell_ospf = cell_ospf.lstrip()
			if ospf_process.match(cell_ospf):
				r = r + 1
				cell_list = cell_ospf.split(' ')
				if cell_list[3] == 'vrf':
					sheet_ospf.cell(r,1).value = cell_list[2]
					sheet_ospf.cell(r,2).value = cell_list[4]
				else:
					sheet_ospf.cell(r,1).value = cell_list[2]
			elif ospf_area.match(cell_ospf):
				cell_list = cell_ospf.split(' ')
				cell_ospf = cell_list[1]
				sheet_ospf.cell(r,3).value = cell_ospf
			elif ospf_network.match(cell_ospf):
				if sheet_ospf.cell(r,4).value == None:
					sheet_ospf.cell(r,4).value = cell_ospf
				else:
					r = r + 1
					sheet_ospf.cell(r,4).value = cell_ospf
					#因为network是很多的，需要进行分行处理
			elif ospf_te.match(cell_ospf):
				cell_list = cell_ospf.split(' ')
				sheet_ospf.cell(r,5).value = cell_ospf
			elif ospf_red.match(cell_ospf):
				cell_list = cell_ospf.split(' ')
				cell_ospf = cell_list[1]
				if sheet_opf.cell(r,6) == None:
					sheet_ospf.cell(r,6).value = cell_ospf
				else:
					r = r + 1
					sheet_ospf.cell(r,6).value =cell_ospf
			elif ospf_module_end.match(cell_ospf):
				n = q + 1
				break
				print("OSPF部分信息完成写入！！！")
			else:
				continue
			#交换机三层口只有vlan，所以基本ospf的配置不会使用接口来配置
	elif vrrp_module.match(cell):
		sheet_vrrp = newwb['vrrp']
		vrrp_count = n
		print('开始进行VRRP信息写入！！！')
		for s in range(vrrp_count,nrows):
			cell_vrrp = con_sheet.cell(s,1).value
			cell_vrrp = cell_vrrp.lstrip()
			if vrrp_intf.match(cell_vrrp):
				t = t + 1
				cell_list = cell_vrrp.split(' ')
				cell_vrrp = cell_list[1]
				sheet_vrrp.cell(t,1).value = cell_vrrp
			elif vrrp_number.match(cell_vrrp):
				cell_list = cell_vrrp.split(' ')
				sheet_vrrp.cell(t,2).value = cell_list[1]
				sheet_vrrp.cell(t,3).value = cell_list[3]
			elif vrrp_priority.match(cell_vrrp):
				cell_list = cell_vrrp.split(' ')
				cell_vrrp = cell_list[3]
				sheet_vrrp.cell(t,4).value = cell_vrrp
			elif vrrp_module_end.match(cell_vrrp):
				n = s + 1
				break
				print("VRRP部分已经完成输出！！！")
			else:
				continue
	elif span_module.match(cell):
		sheet_span = newwb['span']
		span_count = n
		print('开始进行SPAN信息写入！！！')
		for u in range(span_count,nrows):
			cell_span = con_sheet.cell(u,1).value
			cell_span = cell_span.lstrip()
			if span_session_num.match(cell_span):
				v = v + 1
				cell_list = cell_span.split(' ')
				cell_span = cell_list[2]
				sheet_span.cell(v,4).value = cell_span
			elif span_dest_intf.match(cell_span):
				cell_list = cell_span.split(' ')
				sheet_span.cell(v,2).value = cell_list[1]
				sheet_span.cell(v,1).value = cell_list[3]
				#目的只能是物理口

			elif span_source.match(cell_span):
				cell_list = cell_span.split(' ')

				v = v + 1

				sheet_span.cell(v,4).value = cell_list[3]
				sheet_span.cell(v,2).value = cell_list[4]
				sheet_span.cell(v,1).value = cell_list[5] + ' ' + cell_list[6]
				sheet_span.cell(v,3).value = cell_list[8]

			elif span_end_module.match(cell_span):
				n = u + 1
				print('SPAN部分已经输出完毕！！！')
				break
			else:
				pass



	else:
		#print("useless line")
		#这个else在if下面，不在while下面
		#虽然没有匹配但是还是需要继续向下添行的
		continue
		print("n")
else:
	#print(n)
	pass

#这个for循环在中途break了，因为外部的for无法和嵌套的for循环互通
#这个时候，考虑while语句实现。

newwb.save('output.xlsx')
print("成功保存Excel文件！！！")




'''
for line in open("test.txt","r"):
	#读取TXT中的数据，按行输出
	line = line[:-1]
	#删除换行符
	if vrf_module.match(line):
		sheet_vrf = newwb.get_sheet('vrf')
	else:
		break
	if vrf_rd
	sheet_vrf.wtite(a,b,line)

	a = a+1

'''

title = gui.msgbox(msg="已经完成配置文件读取，保存在output.xlsx！！！",title="恭喜完成",ok_button="好的",image="finish.gif")
#msgbox(msg='(Your message goes here)', title=' ', ok_button='OK', image=None, root=None)



'''
#sheet.write(a,b,line)
#xlwt的行、列都是从0开始的


import easygui as g


if g.ccbox(msg = "是否继续输入实际接口物理状态？",choices = ("是的","不用了"),image = "haha2.gif"):
	g.msgbox(msg = "ok go on",image = "haha2.gif")
else:
	g.msgbox("ok qiut")

def Print_Access():  #定义Access模块
	file.write(' interface ' + str(sheet1.cell(a,0).value))
	file.write('\n')
	file.write(' switchport mode access')
	file.write('\n')
	file.write(' switchport access vlan ' + str(int(sheet1.cell(a,1).value)))
	file.write('\n')
	file.write(' $')
	file.write('\n')

def Print_Trunk():   #定义Trunk模块
	file.write(' interface ' + str(sheet1.cell(a,0).value))
	file.write('\n')
	file.write(' switchport mode trunk')
	file.write('\n')
	trunk_vlan = sheet1.cell(a,1).value   #如果是trunk，那应该是一串数字
	if (type(trunk_vlan) == float):    #只有一个vlan那就是数字
		file.write(' switchport trunk vlan ' + str(int(trunk_vlan)))
		file.write('\n')
		file.write('$' + '\n')
	else:
		list1 = trunk_vlan.split(',')  #如果是多个vlan，那就是逗号分隔的字符串，用分隔符隔离成数组
		for X in list1:
			file.write(' switchport trunk vlan ' + X)
			file.write('\n')
		file.write('$')
		file.write('\n')


FinishTime = time.strftime('%Y-%m-%d-%H-%S')#导入系统时间，自定义格式

workbook = xlrd.open_workbook('8900.xlsx')


for x in workbook.sheet_names():#按照表格中的名字来循环
	sheet1 = workbook.sheet_by_name(x)
	nrows = sheet1.nrows
	ncols = sheet1.ncols
	file = open( x + '-' + FinishTime + '.txt','a')
	file .write('conf t' + '\n')
	file.write('switchvlan-configuration' + '\n')

	for a in range(1,nrows):#子循环调用上面两个函数输出脚本
		int_type = str(sheet1.cell(a,2).value)#判断是什么类型的接口
		if int_type == "Access":
	 		Print_Access()
		elif int_type == "Trunk":
	 		Print_Trunk()
		else:
	 		print("wrong Interface Type")#故障感知，而不是程序自动跳出
	file.write('$')
	file.close()
	print('finish output configuartion of ' + x)#每一段都需要回显好监控是否有问题

print('all done')



#msg = g.msgbox("Hello Easy GUI")
tittle = g.msgbox(msg = "所有交换机脚本输出完成！",title = "标题",ok_button="好的",image="haha.gif" )
#最后需要提示脚本运行完成

'''
