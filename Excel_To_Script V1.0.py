#_*_coding=utf-8_*_

from openpyxl import load_workbook
import re
import easygui as gui


'''
目前已经完成的功能模块： 20190522
VRF VLAN LACP Vrrp L2L3port Span

需要添加的模块：
Ospf Bgp MplsL2Vpn MplsL3Vpn

MPLS_te

'''
'''
改动记录：
20190523 vlan部分没有hybrid，需要修改
20190523 发现在Excel里面没有span信息，已经补全
20190523 span部分和ros不一样，需要进行更改

'''



#1.读取Excel里面sheet的数据，按照逻辑进行排列并且进行输出
#2.输出到Excel比较好？？？ 各自有各自的模块，所以，不会互相干扰

wb = load_workbook('output.xlsx')
#打开整理好了的Excel表文件

con_wb = load_workbook('configure.xlsx')
#打开后续存储配置的Excel表

'''
vrf还是错了，每次进ipv4 address-family 都是需要退出的。不然退不出去。
vrf的description 也是可以添加空格的，需要注意一下。
'''
'''
M6000 的配置如图： 表现形式是不一样的。
!<vrf>
ip vrf hd_iptv
  rd 64964:3061000
  route-target import 64964:3061000
  route-target export 64964:3061000
  address-family ipv4
  $
$
'''


def vrf_commands():
	vrf_sheet = wb['vrf']
	#打开vrf的sheet
	#vrf_data 怎么样？？？

	nrows = vrf_sheet.max_row     #行
	ncols = vrf_sheet.max_column  #列
	print(nrows)
	print(ncols)

	vrf_con_sheet = con_wb['vrf_configure']
	#打开用来存储VRF配置的sheet

	i = 2  #原始配置Excel里面的行，从第2行开始

	n = 1

	a = 1
	b = 3

	vrf_con_sheet.cell(n,1).value = '!<vrf>'
	#首行代表模块起始
	n = n + 1


	while i <= nrows:
		#有几行的数据，就说明有几个VRF
		#while a < ncols:
			#有四列数据，需要循环读取，输出,
			#但是在这里只要一行就行了，其他列直接公式输出
			source = vrf_sheet.cell(i,1).value
			line = 'ip vrf ' + source
			print(line)
			vrf_con_sheet.cell(n,1).value = line
			n = n + 1
			#写入第一行，定义VRF
			if vrf_sheet.cell(i,5).value == None:
				pass
			else:
				source = vrf_sheet.cell(i,5).value
				line = 'description ' + source
				vrf_con_sheet.cell(n,1).value = line
				n = n + 1
			#写入第二行，VRF文字描述
			if vrf_sheet.cell(i,2).value == None:
				pass
			else:
				source = vrf_sheet.cell(i,2).value
				line = 'rd ' + source
				vrf_con_sheet.cell(n,1).value = line
				n = n + 1
			#写入第三行，定义RD
			line = 'address-family ipv4'
			vrf_con_sheet.cell(n,1).value = line
			n = n + 1
			#写入第四行，进入address-family
			source = vrf_sheet.cell(i,3).value
			line = 'route-target import ' + source
			vrf_con_sheet.cell(n,1).value = line
			n = n + 1
			#写入第五行，RT，import
			source = vrf_sheet.cell(i,4).value
			line = 'route-target export ' + source
			vrf_con_sheet.cell(n,1).value = line
			n = n + 1 
			#写入第六行，RT export
			vrf_con_sheet.cell(n,1).value = '$'
			n = n + 1
			vrf_con_sheet.cell(n,1).value = '$'
			n = n + 1

			i = i + 1
			#把所有的行写完，需要换下一行
			#注意一下：因为VRF不是在单独的模块里面，所以不需要最后的$来退出模块

	vrf_con_sheet.cell(n,1).value = '!</vrf>'
	n = n + 1



def vlan_commands():
	vlan_sheet = wb['vlan']
	vlan_con_sheet = con_wb['vlan_configure']
	nrows = vlan_sheet.max_row
	ncols = vlan_sheet.max_column
	#print(nrows)
	#print(ncols)

	i = 2  #用作：vlan、vlan-list、接口下vlan配置等的计数工作
	n = 1
	a = 0  #a用来进行trunk的循环
	c = 0  #c用来做hybrid tag的循环
	d = 0  #D用来做hybrid untag的循环
	b = 2  #用来vlan list的循环
	vlan_con_sheet.cell(n,1).value = '!<switchvlan>'
	n = n + 1
	vlan_con_sheet.cell(n,1).value = 'switchvlan-configuration'
	n = n + 1
	#print(n)
	#先写VLAN信息。
	while i <= nrows:

		if vlan_sheet.cell(i,9).value == None:
			pass
			#因为vlan肯定要比总行数小，所以会有空白行。
		else:
			source = vlan_sheet.cell(i,9).value
			line = 'vlan ' + source
			vlan_con_sheet.cell(n,1).value = line
			n = n + 1

			if vlan_sheet.cell(i,10).value == None:
				pass
			else:
				source = vlan_sheet.cell(i,10).value
				line = 'name ' + source
				vlan_con_sheet.cell(n,1).value = line
				n = n + 1
			vlan_con_sheet.cell(n,1).value = '$'
			n = n + 1
		i = i + 1



	#这里是vlan-list
	i = 2
	while i <= nrows:
		if vlan_sheet.cell(i,11).value == None:
			pass
		else:
			source = vlan_sheet.cell(i,11).value
			line = 'list ' + source
			vlan_con_sheet.cell(n,1).value = line
			n = n + 1
			#list下面是没有 美刀符号的
		i = i + 1

	#这里是接口下VLAN配置
	i = 2
	while i <= nrows:
		#print(i)
		#print(vlan_sheet.cell(i,2).value)
		if vlan_sheet.cell(i,2).value == 'trunk':
			source = vlan_sheet.cell(i,1).value
			line = 'interface ' + source
			#print(line)
			vlan_con_sheet.cell(n,1).value = line
			n = n + 1
			vlan_con_sheet.cell(n,1).value = 'switchport mode trunk'
			n = n + 1
			source = vlan_sheet.cell(i,4).value
			vlan_list = source.split(',')
			for a in vlan_list:
				line = 'switchport trunk vlan ' + a
				#print(line)
				vlan_con_sheet.cell(n,1).value = line
				n = n + 1
				#print(n)
			if vlan_sheet.cell(i,5).value == None:
				pass
			else:
				source = vlan_sheet.cell(i,5).value
				line = 'switchport trunk native vlan ' + source
				#print(line)
				vlan_con_sheet.cell(n,1).value = line
				n = n + 1
			vlan_con_sheet.cell(n,1).value = '$'
			n = n + 1
			#print(n)
		elif vlan_sheet.cell(i,2).value == 'hybrid':
			source = vlan_sheet.cell(i,1).value
			line = 'interface ' + source
			vlan_con_sheet.cell(n,1).value = line
			n = n + 1
			vlan_con_sheet.cell(n,1).value = 'switchport mode hybrid'
			n = n + 1

			source = vlan_sheet.cell(i,6).value
			vlan_list = source.split(',')
			for c in vlan_list:
				line = 'switchport hybrid vlan '+ c + ' tag'
				vlan_con_sheet.cell(n,1).value = line
				n = n + 1

			if vlan_sheet.cell(i,7).value == None:
				pass
			else:
				source = vlan_sheet.cell(i,7).value
				vlan_list = source.split(',')
				for d in vlan_list:
					line = 'switchport hybrid vlan ' + d + ' untag'
					vlan_con_sheet.cell(n,1).value = line
					n = n + 1
			vlan_con_sheet.cell(n,1).value = '$'
			n = n + 1

		elif vlan_sheet.cell(i,2).value == 'access':
			source = vlan_sheet.cell(i,1).value
			line = 'interface ' + source
			#print(line)
			vlan_con_sheet.cell(n,1).value = line
			n = n + 1
			source = vlan_sheet.cell(i,3).value
			line = 'switchport access vlan ' + source
			vlan_con_sheet.cell(n,1).value = line
			n = n + 1
			vlan_con_sheet.cell(n,1).value = '$'
			n = n + 1
			#print(n)
		else:
			pass
			#很可能有些接口仅仅是配置了但是没有任何配置,这里应该break掉进入下一行的配置
			#后面还有vlan的名字和list
		i = i + 1
		#print(n)
		#i 这个变量需要进行更迭
	vlan_con_sheet.cell(n,1).value = '$'
	n = n + 1
	#print(n)
	vlan_con_sheet.cell(n,1).value = '!</switchvlan>'

'''
！！！！！！！！！！！！！！！！！！！！！！！！！
  interface smartgroup1
    lacp mode 802.3ad
    lacp load-balance enhance
  $      
  这个还是没有的是load-balance
'''



def lacp_commands():
	lacp_sheet = wb['lacp']
	lacp_con_sheet = con_wb['lacp_configure']
	nrows = lacp_sheet.max_row
	ncols = lacp_sheet.max_column
	i = 2 #用于原始数据的行数
	n = 1 #用于输出配置的Excel行数

	lacp_con_sheet.cell(n,1).value = '!<lacp>'
	n = n + 1
	lacp_con_sheet.cell(n,1).value = 'lacp'
	n = n + 1

	lacp_port = re.compile('smartgroup')



	while i <= nrows:
		if lacp_port.match(lacp_sheet.cell(i,1).value):
			source = lacp_sheet.cell(i,1).value
			print(source)
			line = 'interface ' + source
			lacp_con_sheet.cell(n,1).value = line
			n = n + 1

			source = lacp_sheet.cell(i,4).value
			line = 'lacp mode ' + source
			lacp_con_sheet.cell(n,1).value = line
			n = n + 1

			if lacp_sheet.cell(i,5).value == None:
				pass
			else:
				source = lacp_sheet.cell(i,5).value
				line = 'lacp load-balance ' + source
				lacp_con_sheet.cell(n,1).value = line
				n = n + 1

			lacp_con_sheet.cell(n,1).value = '$'
			n = n + 1

		else:
			source = lacp_sheet.cell(i,1).value
			line = 'interface ' + source
			lacp_con_sheet.cell(n,1).value = line
			n = n + 1

			sg_number = lacp_sheet.cell(i,2).value
			sg_mode = lacp_sheet.cell(i,3).value
			line = 'smartgroup ' + sg_number + ' mode ' + sg_mode
			lacp_con_sheet.cell(n,1).value = line
			n = n + 1

			lacp_con_sheet.cell(n,1).value = '$'
			n = n + 1

		i = i + 1
	lacp_con_sheet.cell(n,1).value = '$'
	n = n + 1
	lacp_con_sheet.cell(n,1).value  = '!</lacp>'
	n = n + 1


def vrrp_commands():
	vrrp_sheet = wb['vrrp']
	vrrp_con_sheet = con_wb['vrrp_configure']
	nrows = vrrp_sheet.max_row
	ncols = vrrp_sheet.max_column

	i = 2
	n = 1

	vrrp_con_sheet.cell(n,1).value = '!<vrrp>'
	n = n + 1
	vrrp_con_sheet.cell(n,1).value = 'vrrp'
	n = n + 1

	while i <= nrows:

		source = vrrp_sheet.cell(i,1).value
		line = 'interface ' + source
		vrrp_con_sheet.cell(n,1).value = line
		n = n + 1

		vrrp_number = str(vrrp_sheet.cell(i,2).value)
		print(vrrp_number)
		vrrp_ip = str(vrrp_sheet.cell(i,3).value)
		print(vrrp_ip)


		line = 'vrrp ' + vrrp_number + ' ipv4 ' + vrrp_ip
		vrrp_con_sheet.cell(n,1).value = line
		n = n + 1

		if vrrp_sheet.cell(i,4).value == None:
			pass
		else:
			vrrp_priority = str(vrrp_sheet.cell(i,4).value)
			line = 'vrrp ' + vrrp_number + ' priority ' + vrrp_priority
			vrrp_con_sheet.cell(n,1).value = line
			n = n + 1

		if vrrp_sheet.cell(i,5).value == None:
			pass
		else:
			line = 'vrrp ' + vrrp_number + ' preempt'
			vrrp_con_sheet.cell(n,1).value = line
			n = n + 1

		vrrp_con_sheet.cell(n,1).value = '$'
		n = n + 1

	i = i + 1

	vrrp_con_sheet.cell(n,1).value = '$'
	n = n + 1
	vrrp_con_sheet.cell(n,1).value = '!</vrrp>'
	n = n +1

'''
接口的no shutdown 写进去了没？
交换机的pw 是在intf模块里面是可以看到的。就是不知道在配置文件里面是不是也是这样。
写进去了，好处是89E三层口没有no shutdown 的问题。
'''

def l2l3port_commands(): # 能不能直接所有的都在一个大的if语句里面
	l2port_sheet = wb['l2port']
	l2port_con_sheet = con_wb['l2port_configure']

	nrows = l2port_sheet.max_row
	ncols = l2port_sheet.max_column

	i = 2
	n = 1

	l2port_con_sheet.cell(n,1).value = '!<if-intf>'
	n = n + 1

	while i <= nrows:

		source = l2port_sheet.cell(i,1).value
		line = 'interface ' + source
		l2port_con_sheet.cell(n,1).value = line
		n = n + 1

		if l2port_sheet.cell(i,2).value == 'L2':
			if l2port_sheet.cell(i,3).value == None:
				pass
			else:
				source = l2port_sheet.cell(i,3).value
				line = source
				l2port_con_sheet.cell(n,1).value = line
				n = n + 1

			if l2port_sheet.cell(i,6).value == None:
				pass
			else:

				source = l2port_sheet.cell(i,6).value
				line = 'description ' + source
				l2port_con_sheet.cell(n,1).value = line
				n = n + 1

			l2port_con_sheet.cell(n,1).value = '$'
			n = n + 1

		elif l2port_sheet.cell(i,2).value == 'L3':

			if l2port_sheet.cell(i,6).value == None:
				pass
			else:
				source = l2port_sheet.cell(i,6).value
				line = 'description ' + source
				l2port_con_sheet.cell(n,1).value = line
				n = n + 1

			if l2port_sheet.cell(i,5).value == None:
				pass
			else:
				source = l2port_sheet.cell(i,5).value
				line = 'ip vrf forwarding ' + source
				l2port_con_sheet.cell(n,1).value = line
				n = n + 1

			if l2port_sheet.cell(i,4).value == None:
				pass
			else:
				source = l2port_sheet.cell(i,4).value
				line = 'ip address ' + source
				l2port_con_sheet.cell(n,1).value = line
				n = n + 1

			l2port_con_sheet.cell(n,1).value = '$'
			n = n + 1


		elif l2port_sheet.cell(i,2).value == 'smartgroup口':
			if l2port_sheet.cell(i,6).value == None:
				pass
			else:
				source = l2port_sheet.cell(i,6).value
				line = 'description ' + source
				l2port_con_sheet.cell(n,1).value = line
				n = n + 1

			l2port_con_sheet.cell(n,1).value = '$'
			n = n + 1

		else:
			#第一，为了防止特殊情况，需要把任然有接口数据的都循环一遍。
			if l2port_sheet.cell(i,3).value == None:
				pass
			else:
				source = l2port_sheet.cell(i,3).value
				line = source
				l2port_con_sheet.cell(n,1).value = line
				n = n + 1

			if l2port_sheet.cell(i,5).value == None:
				pass
			else:
				source = l2port_sheet.cell(i,5).value
				line = 'ip vrf forwarding ' + source
				l2port_con_sheet.cell(n,1).value = line
				n = n + 1

			if l2port_sheet.cell(i,4).value == None:
				pass
			else:
				source = l2port_sheet.cell(i,4).value
				line = 'ip address ' + source
				l2port_con_sheet.cell(n,1).value = line
				n = n + 1

			if l2port_sheet.cell(i,6).value == None:
				pass
			else:
				source = l2port_sheet.cell(i,6).value
				line = 'description ' + source
				l2port_con_sheet.cell(n,1).value = line
				n = n + 1

			l2port_con_sheet.cell(n,1).value = '$'
			n = n + 1

		i = i + 1

	l2port_con_sheet.cell(n,1).value = '!</if-intf>'
	n = n + 1


def span_commands():
	#span应该是需要先定义span session 才能使用span apply的，这里没有办法，没事，刷两遍即可。
	span_sheet = wb['span']
	span_con_sheet = con_wb['span_configure']

	nrows = span_sheet.max_row
	ncols = span_sheet.max_column

	i = 2
	n = 1

	span_con_sheet.cell(n,1).value = '!<monitor>'
	n = n + 1

	while i <= nrows:

		if span_sheet.cell(i,2).value == None:
			pass

		elif span_sheet.cell(i,2).value == 'destination':

			source = span_sheet.cell(i,4).value
			line = 'span session ' + str(source) 

			span_con_sheet.cell(n,1).value = line
			n = n + 1

			source = span_sheet.cell(i,1).value
			line = 'default destination interface ' + source

			span_con_sheet.cell(n,1).value = line
			n = n + 1

			span_con_sheet.cell(n,1).value = '$'
			n = n + 1

		elif span_sheet.cell(i,2).value == 'source':

			intf_name = span_sheet.cell(i,1).value
			span_session_num = span_sheet.cell(i,4).value
			span_source_way = span_sheet.cell(i,3).value

			line = 'span apply session ' + span_session_num + ' source ' + intf_name + ' direction ' + span_source_way

			span_con_sheet.cell(n,1).value = line
			n = n + 1

		else:
			pass

		i = i + 1

	span_con_sheet.cell(n,1).value = '!</monitor>'
	n = n + 1
	print('SPAN部分配置已经生成完毕！！！')


if __name__ == '__main__':
    vrf_commands()
    vlan_commands()
    lacp_commands()
    vrrp_commands()
    l2l3port_commands()
    span_commands()
    con_wb.save('configure_output.xlsx')


#后续是二层接口的描述、三层接口的IP、描述、MTU、VRF

#vlan信息的录入

#lacp可以等等，毕竟不多，简单。
